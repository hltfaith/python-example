from django.test import TestCase

# Create your tests here.
from openpyxl import load_workbook

def xslx_data():

    wb = load_workbook('/home/changhao/PycharmProjects/northcorezh/northcorePro_v1.0/test/模板.xlsx')  #加载excel
    # print(wb.get_sheet_names())  工作表1
    # wb.get_sheet_by_name('工作表1')
    sheet = wb.active

    # print(sheet.max_column)  #查看最大列
    # print(sheet.max_row)     #查看最大行

    # 将excel数据转换字典格式存放
    data = {}
    zy_list = []

    # 获取最大行
    row_line = sheet.max_row + 1

    # 遍历excel所有数据,存放到字典
    for i in range(2, row_line):

        for row in sheet.iter_rows('A{0}:K{1}'.format(i, i)):

            for call in row:
                zy_list.append(call.value)

            data[i] = str(zy_list)
            zy_list.clear()

        i += 1

    # 查看字典
    # for dict in data:
    #     print(data[dict])

    return data