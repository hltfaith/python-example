import openpyxl
import re

wb = openpyxl.load_workbook('account.xlsx')
#print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('account')

a_list = []                     # 存放 excel A1-4098
for i in range(4098):
    i += 1
    ai = ('A' + str(i))
    a_list.append(ai)

h_list = []
for k in range(4098):           # 存放 excel H1-4098
    k += 1
    hk = ('H' + str(k))
    h_list.append(hk)

ah_dic = {}                     # 存放excel A1-4098, H1-4098    ---> 'A1':'H1'
count = 1
while count <= 4098:
    adic = ('A' + str(count))
    hdic = ('H' + str(count))
    ah_dic[adic] = hdic
    count += 1

ab_dic = {}                    # 存放excel A1-4098, B1-4098    ---> 'A1':'B1'
count = 1
while count <= 4098:
    adic = ('A' + str(count))
    bdic = ('B' + str(count))
    ab_dic[adic] = bdic
    count += 1

a_value_list = []
h_value_list = []

def jxs():
    for name in a_list:
        a_value = sheet[name].value
        match = re.search('[a-z]{10}', str(a_value))
        if match:
            a_value_list.append(a_value)
            h_value_list.append(sheet[ah_dic[name]].value)

def cydw():
    for name in a_list:
        a_value = sheet[name].value
        match = re.search('[\d]', str(a_value))
        if match:
            a_value_list.append(a_value)
            h_value_list.append(sheet[ah_dic[name]].value)

def export_user_xlsx():
    ws = openpyxl.Workbook()
    ws1 = ws.active
    dest_filename = 'vpnuser.xlsx'

    num1 = 1
    for name in a_value_list:
        a = ('A' + str(num1))
        ws1[a] = name
        num1 += 1

    num2 = 1
    for value in h_value_list:
        b = ('B' + str(num2))
        ws1[b] = value
        num2 += 1

    ws.save(dest_filename)
    print("VPN用户信息导出成功！！！")


def extract_user_value():

    user_all = {}           # 'name':'value'
    def user_all_data():                    # 数据库全部用户及状态提取
        for name in a_list:
            a_value = sheet[name].value
            user_all[a_value] = sheet[ah_dic[name]].value

    jxs_a_list = []  # 存放 excel A1-1702
    for i in range(1702):
        i += 1
        ai = ('A' + str(i))
        jxs_a_list.append(ai)

    jxs_ab_dic = {}  # 存放excel A1-1702, B1-4098    ---> 'A1':'B1'
    count = 1
    while count <= 1702:
        adic = ('A' + str(count))
        bdic = ('B' + str(count))
        jxs_ab_dic[adic] = bdic
        count += 1

    current_user = {}          # 'name':'value'  存放用户和状态
    current_user_list = []      # 存放重复用户名
    def current_user_data():                # 用户提供经销商用户及状态信息方法
        yhtg = openpyxl.load_workbook('test.xlsx')
        sheet1 = yhtg.get_sheet_by_name('Sheet1')

        aa_list = 0
        cz_list = 0

        for name in jxs_a_list:
            aa_list += 1
            cz_list += 1
            #cz_list = str(cz_list) + '重复账号'
            aa_str = 'AA' + str(aa_list)

            a_value = sheet1[name].value  # xingming
            if a_value == 'aaaaaaaaa':
                current_user[aa_str] = aa_str
                continue

            elif a_value in current_user_list:
                current_user1 = a_value + str(cz_list) + '重复账号'
                current_user[current_user1] = '已重复'
                continue

            else:
                current_user[a_value] = sheet1[jxs_ab_dic[name]].value
                current_user_list.append(sheet1[name].value)
                continue

    exist_user_list = []            # 收集已存在vpn用户
    exist_user_dict = {}            # 收集已存在用户当前状态

    def collection_user_data():
        count11 = 0
        for i in current_user:  # 'name'
            count11 += 1
            count11_str = '黑户' + str(count11)
            if i in user_all.keys():
                if current_user[i] == user_all[i]:
                    #print(i, "       ", current_user[i], "-------------------------->", i, "       ", user_all[i])
                    exist_user_list.append(i)
                    exist_user_dict[i] = user_all[i]

                elif current_user[i] == None:
                    #print(i, "       ", current_user[i], "-------------------------->", i, "       ", user_all[i])
                    exist_user_dict[i] = user_all[i]

                else:
                    #print('\033[1;31m异常error!!!!!!!\033[0m')
                    exist_user_dict[i] = user_all[i]
                    #pass

            elif i == 'None':
                #print(i, "       ", current_user[i], "-------------------------->", i, "       ", user_all[i])
                exist_user_dict[i] = user_all[i]

            else:
                #print(i, "       ", current_user[i], "--------------------------->", "\033[1;31m黑户!!!!!!\033[0m")
                exist_user_dict[i] = '黑户'

    def collection_write_excel():
        ws = openpyxl.Workbook()
        ws1 = ws.active
        dest_filename = 'jxs_user_sum.xlsx'

        num1 = 0
        for name in exist_user_dict:        # 将用户提供跟数据库对比部分 写入excel
            num1 += 1
            c = ('C' + str(num1))
            d = ('D' + str(num1))
            # print(name,"---------------------------->",num1)
            ws1[c] = name
            ws1[d] = exist_user_dict[name]

        num1 = 0
        for name in exist_user_dict:        # 将用户提供跟数据库对比部分 写入excel
            num1 += 1
            c = ('C' + str(num1))
            d = ('D' + str(num1))
            ws1[c] = name
            ws1[d] = exist_user_dict[name]


        not_exist_user_dict = {}
        for name in user_all:               # 将剩余用户存放到字典中
            if name == exist_user_list:
                pass
            else:
                not_exist_user_dict[name] = user_all[name]

        jxs_user_dict = {}
        cydw_user_dict = {}
        for name in not_exist_user_dict:                    # 将经销商用户和成员单位用户分类
            name_re = re.search('[\d]', str(name))
            if name_re:
                cydw_user_dict[name] = not_exist_user_dict[name]
            else:
                jxs_user_dict[name] = not_exist_user_dict[name]

        num2 = 1703
        for name in jxs_user_dict:    # 将经销商用户写入C、D列
            c = ('C' + str(num2))
            d = ('D' + str(num2))
            ws1[c] = name
            ws1[d] = jxs_user_dict[name]
            num2 += 1

        num3 = 1
        for name in cydw_user_dict:  # 将成员单位用户写入E、F列
            e = ('E' + str(num3))
            f = ('F' + str(num3))
            ws1[e] = name
            ws1[f] = cydw_user_dict[name]
            num3 += 1

        ws.save(dest_filename)
        print("VPN用户信息导出成功！！！")

    def start():
        user_all_data()
        current_user_data()

        print('VPN用户数据从excel提取成功!!!!!')
        collection_user_data()
        print('VPN用户数据与数据库中数据收集成功!!!!!')
        collection_write_excel()

    return start()

def export_cgdw_xlsx():

    cydw_user = []
    cydw_user_status = []
    other_cydw = []
    other_cydw_status = []

    def select_cydw():
        wb = openpyxl.load_workbook('cydw.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')

        a_list = []  # 存放 excel A1-2317
        for i in range(2317):
            i += 1
            ai = ('A' + str(i))
            a_list.append(ai)

        b_list = []
        for k in range(2317):  # 存放 excel B1-2317
            k += 1
            bk = ('B' + str(k))
            b_list.append(bk)

        ab_dic = {}  # 存放excel A1-4098, H1-4098    ---> 'A1':'H1'
        count = 1
        while count <= 2317:
            adic = ('A' + str(count))
            bdic = ('B' + str(count))
            ab_dic[adic] = bdic
            count += 1

        for name in a_list:
            a_value = sheet[name].value
            match = re.findall('\D', str(a_value))

            if match:
                cydw_user.append(a_value)
                cydw_user_status.append(sheet[ab_dic[name]].value)   # B1 A1
            else:
                other_cydw.append(a_value)
                other_cydw_status.append(sheet[ab_dic[name]].value)

    select_cydw()

    def export_cydw():

        ws = openpyxl.Workbook()
        ws1 = ws.active
        dest_filename = 'cydw1.xlsx'

        cydw_num = 1                    # A列 成员单位用户名
        for name1 in cydw_user:
            a = ('A' + str(cydw_num))
            ws1[a] = name1
            cydw_num += 1

        cydw_status = 1                 # B列 成员单位用户状态
        for value1 in cydw_user_status:
            b = ('B' + str(cydw_status))
            ws1[b] = value1
            cydw_status += 1

        other_cydw_num = 1              # C列 成员单位用户名
        for name2 in other_cydw:
            a = ('C' + str(other_cydw_num))
            ws1[a] = name2
            other_cydw_num += 1

        other_cydw_value = 1            # D列 成员单位用户状态
        for value2 in other_cydw_status:
            b = ('D' + str(other_cydw_value))
            ws1[b] = value2
            other_cydw_value += 1

        ws.save(dest_filename)
        print("成员单位用户信息导出成功！！！ 新文件 ---> %s" % dest_filename)

    export_cydw()

def diff_cms_user():

    def select_cms():
        wb = openpyxl.load_workbook('cmsuser.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')

        a_list = []  # 存放 excel A1-2005
        for i in range(2005):
            i += 1
            ai = ('A' + str(i))
            a_list.append(ai)

        ab_dic = {}  # 存放excel A1-2005, B1-2005    ---> 'A1':'B1'
        count = 1
        while count <= 2005:
            adic = ('A' + str(count))
            bdic = ('B' + str(count))
            ab_dic[adic] = bdic
            count += 1

        # {'公司':{'changhao':'status'}}
        # cms 2005
        # cms_tongji 1771

        # 1. 现有经销商与总表对比

        zong_cms_dict = {}                  # 账号，公司名
        for name in a_list:
            a_value = sheet[name].value
            b_value = sheet[ab_dic[name]].value
            zong_cms_dict[a_value] = b_value

        current_cms_dict = {}           # 账号，状态
        def diff_current_cms():
            wb = openpyxl.load_workbook('jryg.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')

            a_list = []  # 存放 excel A1-1771
            for i in range(1771):
                i += 1
                ai = ('A' + str(i))
                a_list.append(ai)

            ab_dic = {}  # 存放excel A1-1771, B1-1771    ---> 'A1':'B1'
            count = 1
            while count <= 1771:
                adic = ('A' + str(count))
                bdic = ('B' + str(count))
                ab_dic[adic] = bdic
                count += 1

            for name in a_list:
                a_value = sheet[name].value
                b_value = sheet[ab_dic[name]].value    # 状态
                current_cms_dict[a_value] = b_value

        diff_current_cms()

        # 1.当前用户
        comp_user = {}          # {'公司名称':'账号'}
        not_found_user = {}     # {'账号': '状态'}
        for i in current_cms_dict:         # 账号
            if i in zong_cms_dict.keys():  # 账号
                comp_user[zong_cms_dict[i]] = i  # 公司名 账号
            else:
                not_found_user[i] = current_cms_dict[i]

        # 1. 将总表剩余的用户过滤出来
        zong_free_cms = {}  # {'公司名称':'账号'}
        for i in zong_cms_dict:   # 账号
            if i in comp_user.values():
                pass
            else:
                zong_free_cms[zong_cms_dict[i]] = i # 公司名, 账号

        def export_cms():

            def export_current_cms():       #导出当前cms用户的excel数据
                ws = openpyxl.Workbook()
                ws1 = ws.active
                dest_filename = 'cureent_cms1.xlsx'

                # 写入当前cms用户格式  公司名 账号 状态

                cydw_num = 1  # A列 公司名
                for name1 in comp_user:
                    a = ('A' + str(cydw_num))
                    ws1[a] = name1
                    cydw_num += 1

                cydw_status = 1  # B列 账号
                for value1 in comp_user.values():
                    b = ('B' + str(cydw_status))
                    ws1[b] = value1
                    cydw_status += 1

                other_cydw_num = 1  # C列 状态
                for name2 in comp_user:     # 公司名称
                    a = ('C' + str(other_cydw_num))
                    ws1[a] = current_cms_dict[comp_user[name2]]
                    other_cydw_num += 1

                other_cydw_value = 1  # D列 成员单位用户状态
                for value2 in not_found_user:
                    b = ('D' + str(other_cydw_value))
                    ws1[b] = value2
                    other_cydw_value += 1

                other_cydw_value1 = 1  # E列 成员单位用户状态
                for value2 in not_found_user.values():
                    b = ('E' + str(other_cydw_value1))
                    ws1[b] = value2
                    other_cydw_value1 += 1

                ws.save(dest_filename)
                print("当前cms用户信息导出成功！！！ 新文件 ---> %s" % dest_filename)
            export_current_cms()

            def export_free_cms():          #导出总表与当前cms用户比对后的数据
                ws = openpyxl.Workbook()
                ws1 = ws.active
                dest_filename = 'zong_free_cms1.xlsx'

                cydw_num = 1  # A列 公司名
                for name1 in zong_free_cms:
                    a = ('A' + str(cydw_num))
                    ws1[a] = name1
                    cydw_num += 1

                cydw_status = 1  # B列 账号
                for value1 in zong_free_cms.values():
                    b = ('B' + str(cydw_status))
                    ws1[b] = value1
                    cydw_status += 1

                ws.save(dest_filename)
                print("总表剩余cms用户信息导出成功！！！ 新文件 ---> %s" % dest_filename)
            export_free_cms()

        export_cms()

    select_cms()



if __name__ == '__main__':

    menu = ('''
    **********vpn用户授权信息自动化导出***********
        1.导出 --> 经销商用户
        2.导出 --> 成员单位用户
        3.数据库与当前数据对比
        4.成员单位用户导出
        5.CMS账户与公司名称对比
    ''')

    while True:
        print(menu)
        choice = int(input("请选择编号>>: ").strip())
        if choice == 1:
            jxs()
            export_user_xlsx()
            break

        elif choice == 2:
            cydw()
            export_user_xlsx()
            break

        elif choice == 3:
            extract_user_value()
            break

        elif choice == 4:
            export_cgdw_xlsx()
            break

        elif choice == 5:
            diff_cms_user()
            break

        else:
            print("输入有误！请重新输入.....")
            continue
